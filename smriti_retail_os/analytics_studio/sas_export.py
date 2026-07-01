# -*- coding: utf-8 -*-
#
# @file: smriti_retail_os/analytics_studio/sas_export.py
# @description: SMRITI Analytics Studio — Server-side Export Engine
#               Excel via openpyxl (with branding, formatting, merged cells).
#               PDF via HTML/print trigger (browser-based).
# @author: Jawahar R. Mallah
# @version: 1.0.0
#

import frappe
import json
import io
from frappe.utils import flt, cint, nowdate, today


def generate_excel_bytes(report_key, rows, columns, filters=None, view_name=None):
    """
    Generates a branded XLSX file using openpyxl.
    Returns raw bytes suitable for HTTP response.
    
    Features:
    - SMRITI branded header row
    - Auto-column widths
    - Number formatting (Currency, Float, Percent, Int)
    - Frozen top rows (header + title)
    - Conditional color fills for status fields
    - Grand total row at bottom
    - Sheet metadata (report name, generated date, filters applied)
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.throw("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_key[:31]  # Excel sheet name limit

    # ── Style Definitions ──
    NAVY = "1A2B5C"
    BLUE = "2563EB"
    LIGHT_BLUE = "EFF6FF"
    GRAY = "F8FAFC"
    WHITE = "FFFFFF"
    SUCCESS_FILL = "D1FAE5"
    DANGER_FILL = "FEE2E2"
    WARNING_FILL = "FEF3C7"
    TOTAL_FILL = "1E293B"

    def make_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def make_font(bold=False, size=10, color="000000"):
        return Font(bold=bold, size=size, color=color)

    def make_border():
        thin = Side(style="thin", color="E2E8F0")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ── Row 1: SMRITI Branding ──
    num_cols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "SMRITI Analytics Studio"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = make_fill(NAVY)
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 24

    # ── Row 2: Report Info ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(num_cols, 1))
    info_parts = [f"Report: {report_key.replace('_', ' ').title()}"]
    info_parts.append(f"Generated: {today()}")
    if view_name:
        info_parts.append(f"View: {view_name}")
    info_cell = ws.cell(row=2, column=1)
    info_cell.value = " | ".join(info_parts)
    info_cell.font = Font(italic=True, size=9, color=WHITE)
    info_cell.fill = make_fill(BLUE)
    info_cell.alignment = center_align
    ws.row_dimensions[2].height = 16

    # ── Row 3: Blank separator ──
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column Headers ──
    header_row = 4
    col_fieldtype_map = {}
    col_label_map = {}

    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci)
        label = col.get("label", col.get("fieldname", f"Col{ci}"))
        cell.value = label
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = make_fill("334155")
        cell.alignment = center_align
        cell.border = make_border()
        ws.row_dimensions[header_row].height = 20

        col_fieldtype_map[ci] = col.get("fieldtype", "Data")
        col_label_map[ci] = label

    # ── Data Rows ──
    NUMERIC_TYPES = ("Currency", "Float", "Int", "Percent")
    DATE_TYPES = ("Date", "Datetime")

    for ri, row_data in enumerate(rows, start=header_row + 1):
        row_fill = make_fill(WHITE) if ri % 2 == 1 else make_fill(GRAY)

        for ci, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname", "")
            fieldtype = col.get("fieldtype", "Data")
            raw_val = row_data.get(fieldname)
            cell = ws.cell(row=ri, column=ci)

            if fieldtype in NUMERIC_TYPES and raw_val is not None:
                cell.value = flt(raw_val)
                if fieldtype == "Currency":
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = right_align
                elif fieldtype == "Percent":
                    cell.number_format = '0.00"%"'
                    cell.alignment = right_align
                elif fieldtype == "Float":
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                elif fieldtype == "Int":
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
            elif fieldtype in DATE_TYPES and raw_val:
                cell.value = str(raw_val)
                cell.alignment = center_align
            else:
                cell.value = str(raw_val) if raw_val is not None else ""
                cell.alignment = left_align

            # Conditional fills based on value patterns
            val_str = str(raw_val or "").lower()
            if "out of stock" in val_str:
                cell.fill = make_fill(DANGER_FILL)
            elif "low stock" in val_str:
                cell.fill = make_fill(WARNING_FILL)
            elif "in stock" in val_str:
                cell.fill = make_fill(SUCCESS_FILL)
            else:
                cell.fill = row_fill

            cell.border = make_border()

    # ── Grand Total Row ──
    from smriti_retail_os.analytics_studio.sas_service import compute_grand_totals
    grand_totals = compute_grand_totals(rows, columns)

    if grand_totals and rows:
        total_row_idx = header_row + len(rows) + 1

        # Label cell
        first_col_cell = ws.cell(row=total_row_idx, column=1)
        first_col_cell.value = "GRAND TOTAL"
        first_col_cell.font = Font(bold=True, size=10, color=WHITE)
        first_col_cell.fill = make_fill(TOTAL_FILL)
        first_col_cell.alignment = left_align
        first_col_cell.border = make_border()

        for ci, col in enumerate(columns, start=1):
            if ci == 1:
                continue
            fieldname = col.get("fieldname", "")
            fieldtype = col.get("fieldtype", "Data")
            cell = ws.cell(row=total_row_idx, column=ci)
            cell.fill = make_fill(TOTAL_FILL)
            cell.border = make_border()

            if fieldtype in NUMERIC_TYPES and fieldname in grand_totals:
                cell.value = flt(grand_totals[fieldname])
                cell.font = Font(bold=True, size=10, color=WHITE)
                if fieldtype == "Currency":
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = right_align
                else:
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
            else:
                cell.value = ""

        ws.row_dimensions[total_row_idx].height = 18

    # ── Auto Column Widths ──
    for ci, col in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        label = col.get("label", col.get("fieldname", ""))
        max_len = max(len(str(label)), 10)

        for ri in range(header_row + 1, header_row + min(len(rows), 100) + 1):
            cell_val = ws.cell(row=ri, column=ci).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Freeze Panes (freeze rows 1-4 + first column) ──
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # ── Metadata Sheet ──
    meta_ws = wb.create_sheet("Report Info")
    meta_ws.column_dimensions["A"].width = 25
    meta_ws.column_dimensions["B"].width = 50

    meta_rows = [
        ("Report Key", report_key),
        ("Report Name", report_key.replace("_", " ").title()),
        ("Generated On", str(frappe.utils.now())),
        ("Generated By", frappe.session.user),
        ("Total Rows", len(rows)),
    ]
    if filters:
        for k, v in (filters if isinstance(filters, dict) else {}).items():
            if v:
                meta_rows.append((k.replace("_", " ").title(), str(v)))

    for mi, (key, val) in enumerate(meta_rows, start=1):
        meta_ws.cell(row=mi, column=1).value = key
        meta_ws.cell(row=mi, column=1).font = Font(bold=True)
        meta_ws.cell(row=mi, column=2).value = val

    # ── Return bytes ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def stream_excel_response(report_key, rows, columns, filters=None, view_name=None):
    """
    Streams the Excel file as an HTTP response.
    Called from the whitelisted API endpoint.
    """
    xlsx_bytes = generate_excel_bytes(report_key, rows, columns, filters, view_name)
    filename = f"smriti_{report_key}_{today()}.xlsx"

    frappe.local.response.update({
        "type": "binary",
        "filecontent": xlsx_bytes,
        "filename": filename,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })
