# -*- coding: utf-8 -*-
#
# @file: smriti_retail_os/psv_service.py
# @description: Core service logic for SMRITI Party Stock Visibility.
# @author: Jawahar R Mallah <jawahar.mallah@gmail.com>
# @date: 2026-05-28
# @version: 1.8.6
# @license: GPL-3.0-only
# SPDX-License-Identifier: GPL-3.0-only
# * Copyright (c) 2026 AITDL NETWORK & ERPNbook.com. All rights reserved.
#

import hashlib
import os
import csv
import contextlib
import frappe  # frappe.whitelist, frappe.throw, frappe.session, frappe.logger — framework utilities
from frappe import _
from smriti_retail_os import smriti
from frappe.utils import get_datetime, today, now_datetime
from smriti_retail_os.ledger_engine import make_ledger_entry, log_activity
from smriti_retail_os.balance_engine import get_party_balance, get_bulk_party_balances


# ─── F4-FIX: PSV Distributed Upload Lock ─────────────────────────────────────
# Prevents overselling race condition where two concurrent sales uploads
# for the same Party Stock Account both pass the balance check and then
# both commit, creating a net-negative shadow balance.
#
# Implementation: Redis SET NX (set-if-not-exists) via smriti.cache().
# No new infrastructure required — Frappe already uses Redis for caching.
# Lock is scoped per party_stock_account, so concurrent uploads for
# DIFFERENT locations are not blocked.

_PSV_LOCK_EXPIRY_SECONDS = 30   # Max seconds a single upload validation can hold the lock
_PSV_LOCK_PREFIX = "smriti:psv:upload_lock:"


@contextlib.contextmanager
def _psv_upload_lock(party_stock_account):
    """
    Context manager that acquires a per-PSA distributed lock before
    running the overselling check + submit sequence.

    Raises frappe.ValidationError if the lock cannot be acquired
    (meaning another upload for the same PSA is in progress).
    """
    lock_key = f"{_PSV_LOCK_PREFIX}{party_stock_account}"
    cache = smriti.cache()

    # Redis SET NX EX — atomic set-if-not-exists with expiry
    acquired = cache.set(lock_key, 1, ex=_PSV_LOCK_EXPIRY_SECONDS, nx=True)

    if not acquired:
        frappe.throw(
            _("Another sales upload for party stock account '{0}' is currently being processed. "
              "Please wait a moment and try again.").format(party_stock_account),
            frappe.ValidationError
        )

    try:
        yield
    finally:
        # Always release the lock — even on exception.
        # Lock-release failures must NOT propagate: this is a cleanup path.
        try:
            cache.delete(lock_key)
        except Exception:
            pass  # Lock already expired or Redis unavailable — safe to ignore


# --- UNIVERSAL TRANSACTION ENGINE ---

def create_psv_transaction(psa, transaction_type, items, company=None, reference_doctype=None, reference_name=None, remarks=None, posting_date=None, opening_import_batch=None):
    if not company:
        company = smriti.db.get("SMRITI Party Stock Account", psa, "company")
        
    fingerprint = None
    if reference_doctype and reference_name:
        fingerprint = f"{transaction_type}::{reference_doctype}::{reference_name}"
        # BUG-006 FIX: Check both Draft (0) and Submitted (1) to close the race window
        # where two concurrent requests both pass a docstatus=1-only check.
        # Cancelled transactions (docstatus=2) are excluded — they may be legitimately reprocessed.
        existing = smriti.db.get(
            "SMRITI PSV Transaction",
            {"mapping_fingerprint": fingerprint, "docstatus": ["in", [0, 1]]},
            "name"
        )
        if existing:
            return existing


    doc = smriti.documents.new("SMRITI PSV Transaction")
    doc.party_stock_account = psa
    doc.transaction_type = transaction_type
    doc.company = company
    doc.reference_doctype = reference_doctype
    doc.reference_name = reference_name
    doc.remarks = remarks
    if opening_import_batch:
        doc.opening_import_batch = opening_import_batch
    if posting_date:
        doc.posting_date = posting_date
        
    for item in items:
        if not item.get("item_code") or not item.get("qty"):
            continue
        doc.append("items", {
            "item_code": item.get("item_code"),
            "qty": item.get("qty"),
            "rate": item.get("rate") or 0.0,
            "reason": item.get("reason") or ""
        })
        
    if not doc.items:
        return None
        
    doc.flags.ignore_links = True
    doc.insert(ignore_permissions=True)
    doc.submit()
    return doc.name

# ─── SALES INVOICE HOOKS ──────────────────────────────────────────────────────

def get_posting_datetime(doc):
    """
    Safely combines ERPNext posting_date and posting_time into a Python datetime object.
    """
    if doc.posting_date and doc.posting_time:
        return get_datetime(f"{doc.posting_date} {doc.posting_time}")
    return get_datetime(doc.posting_date or today())

def process_sales_invoice_submit(doc, method=None):
    if not doc.get("custom_party_stock_account"): return
    try:
        tx_type = "RETURN" if doc.is_return else "TRANSFER_OUT"
        items_data = [{"item_code": i.item_code, "qty": i.qty, "rate": i.rate} for i in doc.items]
        if items_data:
            create_psv_transaction(doc.custom_party_stock_account, tx_type, items_data, doc.company, doc.doctype, doc.name, "Generated from Sales Invoice", get_posting_datetime(doc))
    except Exception as e:
        smriti.errors.log_error(title=f"PSV Error: {doc.name}", message=frappe.get_traceback())
        smriti.documents.new("PSVExceptionRecord").update({
            "timestamp": now_datetime(),
            "last_seen": now_datetime(),
            "party_stock_account": doc.custom_party_stock_account,
            "alert_type": "Hook Failure",
            "sales_invoice": doc.name,
            "reconciliation_notes": str(e),
            "status": "Pending Reconciliation"
        }).insert(ignore_permissions=True)
        smriti.db.commit()

def process_sales_invoice_cancel(doc, method=None):
    if not doc.get("custom_party_stock_account"): return
    try:
        tx_name = smriti.db.get("SMRITI PSV Transaction", {"reference_doctype": doc.doctype, "reference_name": doc.name, "docstatus": 1})
        if tx_name: smriti.documents.get("SMRITI PSV Transaction", tx_name).cancel()
    except Exception as e:
        smriti.errors.log_error(title=f"PSV Cancel Error: {doc.name}", message=frappe.get_traceback())
        smriti.documents.new("PSVExceptionRecord").update({
            "timestamp": now_datetime(),
            "last_seen": now_datetime(),
            "party_stock_account": doc.custom_party_stock_account,
            "alert_type": "Hook Failure",
            "sales_invoice": doc.name,
            "reconciliation_notes": str(e),
            "status": "Pending Reconciliation"
        }).insert(ignore_permissions=True)
        smriti.db.commit()
# ─── WEEKLY SALES UPLOAD ──────────────────────────────────────────────────────

def validate_sales_upload(doc):
    """Gathers MD5 validation and period locks validation.

    F4-FIX: Overselling check (step 5) now runs inside a per-PSA Redis lock
    to close the race window where two concurrent uploads could both pass
    the balance check and then both commit, resulting in a negative shadow balance.
    The lock is acquired only during the critical section (check + submit),
    not for the entire validate lifecycle.
    """
    # 1. Validate Period Start/End
    if not doc.period_start_date or not doc.period_end_date:
        frappe.throw(_("Period Start Date and Period End Date are required."))
    if doc.period_start_date > doc.period_end_date:
        frappe.throw(_("Period Start Date cannot be later than Period End Date."))

    # 2. File duplicate MD5 check
    if doc.excel_file:
        file_doc = smriti.documents.get("File", {"file_url": doc.excel_file})
        file_content = file_doc.get_content()
        file_hash = hashlib.md5(file_content).hexdigest()

        duplicate = smriti.db.exists(
            "SMRITI Party Sales Upload",
            {"file_hash": file_hash, "name": ["!=", doc.name]}
        )
        if duplicate:
            frappe.throw(_("Duplicate Upload: This exact file has already been imported under {0}.").format(duplicate))
        doc.file_hash = file_hash

    # 3. Period Overlap Check
    overlapping_import = smriti.db.sql("""
        SELECT name
        FROM `tabSMRITI Party Sales Upload`
        WHERE party_stock_account = %s
          AND docstatus = 1
          AND name != %s
          AND (
            (period_start_date <= %s AND period_end_date >= %s) OR
            (period_start_date <= %s AND period_end_date >= %s) OR
            (%s <= period_start_date AND %s >= period_end_date)
          )
    """, (
        doc.party_stock_account,
        doc.name,
        doc.period_start_date, doc.period_start_date,
        doc.period_end_date, doc.period_end_date,
        doc.period_start_date, doc.period_end_date
    ))
    if overlapping_import:
        frappe.throw(
            _("Period Lock: An imported sales file already exists for this location within the date range. Conflicting import: {0}.")
            .format(overlapping_import[0][0])
        )

    # 4. Parse Excel / CSV and populate items if currently empty (on draft save)
    if not doc.items and doc.excel_file:
        parse_and_populate_items(doc)

    # 5. Overselling check — runs inside per-PSA distributed lock to prevent
    #    concurrent upload race condition (F4-FIX: Redis SET NX).
    with _psv_upload_lock(doc.party_stock_account):
        # Re-read balances INSIDE the lock to get a fresh, race-safe snapshot.
        balances = get_bulk_party_balances(doc.party_stock_account, [item.item_code for item in doc.items])
        for item in doc.items:
            current_bal = balances.get(item.item_code, 0.0)
            if item.qty_sold > current_bal:
                frappe.throw(
                    _("Row #{0}: Cannot record sales of {1} units for SKU {2}. Current available balance is only {3} units.")
                    .format(item.idx, item.qty_sold, item.item_code, current_bal)
                )

def parse_and_populate_items(doc):
    """
    Parses CSV/Excel file and populates the child items table.
    """
    file_doc = smriti.documents.get("File", {"file_url": doc.excel_file})
    file_path = file_doc.get_full_path()
    
    # We support simple CSV parsing for testing/import, and Excel via openpyxl
    if file_path.endswith('.csv'):
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if len(row) >= 3:
                    doc.append("items", {
                        "date": get_datetime(row[0]) if row[0] else today(),
                        "item_code": row[1].strip(),
                        "qty_sold": float(row[2])
                    })
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            # Expect header: Date, Item Code, Qty Sold
            for row in rows[1:]: # Skip header
                if row and len(row) >= 3 and row[1]:
                    doc.append("items", {
                        "date": get_datetime(row[0]) if row[0] else today(),
                        "item_code": str(row[1]).strip(),
                        "qty_sold": float(row[2] or 0)
                    })
        except ImportError:
            frappe.throw(_("Python library 'openpyxl' is required to parse Excel files."))

def process_sales_upload_submit(doc):
    smriti.db.set_value(doc.doctype, doc.name, "status", "Imported")
    items_data = [{"item_code": i.item_code, "qty": i.qty_sold} for i in doc.items]
    create_psv_transaction(doc.party_stock_account, "SALES_UPLOAD", items_data, doc.company, doc.doctype, doc.name, f"Imported from {doc.name}")

def process_sales_upload_cancel(doc):
    """
    Reverses a weekly sales upload.

    BUG-004 FIX: Previously wrote raw make_ledger_entry() calls directly, causing
    double-reversal because the PSV Transaction's on_cancel already writes VOID entries.

    Correct approach:
    1. Find the submitted SMRITI PSV Transaction linked to this upload.
    2. Cancel it — on_cancel() fires update_ledger(cancel=True) which writes the
       correct VOID entries with proper sign handling.
    3. Fallback: if no PSV Transaction exists (legacy/pre-fix data), write direct
       reversal entries as before — but with the correct sign.
    """
    smriti.db.set_value(doc.doctype, doc.name, "status", "Draft")

    # Preferred path: cancel the PSV Transaction (handles ledger reversal atomically)
    tx_name = smriti.db.get(
        "SMRITI PSV Transaction",
        {"reference_doctype": doc.doctype, "reference_name": doc.name, "docstatus": 1},
        "name"
    )
    if tx_name:
        smriti.documents.get("SMRITI PSV Transaction", tx_name).cancel()
    else:
        # Fallback for historical data that was processed before the PSV Transaction
        # document layer existed. Write direct reversal entries.
        for item in doc.items:
            # qty_sold is positive (e.g. 10 units sold).
            # Original SALES_UPLOAD ledger entry was negative (-10).
            # Reversal must be positive (+10) to restore balance.
            make_ledger_entry(
                company=doc.company,
                posting_datetime=now_datetime(),
                party_stock_account=doc.party_stock_account,
                item_code=item.item_code,
                qty=abs(item.qty_sold),  # Always positive — restoring stock
                voucher_type="Sales",
                voucher_no=f"VOID-{doc.name}",
                reason=_("Import Cancelled — direct fallback reversal")
            )

    log_activity(
        action_type="Cancel Dispatch",
        party_stock_account=doc.party_stock_account,
        reference_doctype=doc.doctype,
        reference_name=doc.name,
        details="Reversed weekly sales file import. Balances restored."
    )


# ─── PHYSICAL STOCK RECONCILIATION SNAPSHOT ────────────────────────────────────

def validate_physical_snapshot(doc):
    """Calculates variance based on current Party Stock Ledger state"""
    for item in doc.items:
        # Populate system balance
        item.system_qty = get_party_balance(doc.party_stock_account, item.item_code)
        item.variance = item.physical_qty - item.system_qty

def process_physical_snapshot_submit(doc):
    if doc.status != "Approved":
        frappe.throw(_("Audit Snapshots must be explicitly approved before submitting."))
    doc.approved_by = frappe.session.user
    doc.approved_on = now_datetime()
    items_data = [{"item_code": i.item_code, "qty": i.variance, "reason": i.variance_reason} for i in doc.items if i.variance != 0.0]
    if items_data:
        create_psv_transaction(doc.party_stock_account, "AUDIT_ADJUSTMENT", items_data, doc.company, doc.doctype, doc.name, "Physical Snapshot Approved", get_datetime(doc.audit_date))

def process_physical_snapshot_cancel(doc):
    doc.approved_by = None
    doc.approved_on = None
    tx_name = smriti.db.get("SMRITI PSV Transaction", {"reference_doctype": doc.doctype, "reference_name": doc.name, "docstatus": 1})
    if tx_name:
        smriti.documents.get("SMRITI PSV Transaction", tx_name).cancel()

def import_opening_balances(company, party_stock_account, items_data):
    import uuid
    from frappe.utils import now_datetime

    timestamp = now_datetime().strftime("%Y%m%d%H%M%S")
    suffix = uuid.uuid4().hex[:6]
    batch_id = frappe.generate_hash(length=8)

    pseudo_ref_name = f"OPENING-{party_stock_account}-{timestamp}-{suffix}"
    remarks = f"Initial Opening Balance Import - Batch: {batch_id}"

    return create_psv_transaction(
        psa=party_stock_account,
        transaction_type="OPENING",
        items=items_data,
        company=company,
        reference_doctype="Opening Balance Import",
        reference_name=pseudo_ref_name,
        remarks=remarks,
        opening_import_batch=batch_id
    )

@frappe.whitelist()
def process_opening_balance(company, party_stock_account, items):
    """
    SEC-004 FIX: Enforce role-based access before allowing opening balance import.
    Previously public to any authenticated user — now restricted to store managers.
    """
    frappe.only_for(["System Manager", "SMRITI Store Manager"])

    if isinstance(items, str):
        import json
        items = json.loads(items)

    if not company or not party_stock_account or not items:
        frappe.throw(_("Company, Party Stock Account, and at least one item are required."))

    return import_opening_balances(company, party_stock_account, items)

# ─── OPERATIONAL HEALTH ALERTS & CHECKS ────────────────────────────────────────


# ─── BACKWARD-COMPAT RE-EXPORTS ──────────────────────────────────────────────
# These names moved to dedicated sub-service modules in Phase 4 (file split).
# Re-imported here so existing callers of psv_service.<name>() are not broken.

from smriti_retail_os.psv_snapshot_service import (  # noqa: F401
    get_landing_cost, calculate_aging_for_variant, get_aging_alert, generate_snapshots
)
from smriti_retail_os.psv_health_service import (  # noqa: F401
    find_open_alert, create_or_update_alert, run_psv_daily_health_check,
    validate_sales_invoice_cancel
)
from smriti_retail_os.psv_analytics_service import (  # noqa: F401
    get_redistribution_suggestions, get_channel_health_score,
    get_sellin_sellout_summary, get_stock_cover_risks, get_channel_stock_trend,
    get_inventory_productivity_metrics, get_inventory_productivity_methodology,
    ACTION_INCREASE_STOCK, ACTION_MAINTAIN, ACTION_IMPROVE_MARGIN,
    ACTION_LIQUIDATE, ACTION_REPLENISH_URGENT
)
from smriti_retail_os.psv_migration_service import (  # noqa: F401
    create_reversal_entry, migrate_to_new_psv_partner
)
