# -*- coding: utf-8 -*-
#
# @file: smriti_retail_os/utils/invoice_utils.py
# @description: Platform-level shared invoice utility functions
# @author: Jawahar R Mallah <jawahar.mallah@gmail.com>
# @date: 2026-06-04
# @version: 1.0.0
# @license: MIT
#

import frappe
from frappe.utils import flt
import csv
import io

COL_ALIASES = {
    "barcode":  ["BARCODE", "BARCODE NO", "BARCODE_NO", "EAN", "UPC"],
    "price":    ["PRICE", "SELLING PRICE", "RATE", "MRP", "SELL PRICE"],
    "discount": ["DISCOUNT", "DISC", "DISC %", "DISCOUNT %"],
    "qty":      ["QTY", "QUANTITY", "PCS", "PIECES"],
    "tax":      ["GST", "GST %", "TAX", "TAX %", "GST PERCENT"],
}


def get_barcode_candidates(barcode):
    """
    Normalizes a barcode string and generates potential candidates
    (e.g., stripping trailing .0, stripping leading zeros, EAN/UPC padding/stripping).
    """
    if not barcode:
        return []
    barcode = str(barcode).strip()
    
    # Remove trailing ".0" if it's a float-looking string
    if barcode.endswith(".0") and barcode[:-2].isdigit():
        barcode = barcode[:-2]
        
    candidates = [barcode]
    
    # If the barcode contains digits only
    if barcode.isdigit():
        # Candidate 1: without leading zeros (e.g., "07007..." -> "7007...")
        stripped = barcode.lstrip('0')
        if stripped and stripped not in candidates:
            candidates.append(stripped)
            
        # Candidate 2: EAN-13 padding (12 digits -> 13 digits with leading 0)
        if len(barcode) == 12:
            candidates.append("0" + barcode)
            
        # Candidate 3: GTIN-14 padding (13 digits -> 14 digits with leading 0)
        if len(barcode) == 13:
            candidates.append("0" + barcode)
            
        # Candidate 4: If barcode scanned was 14 digits starting with '0', also try the 13 digit EAN-13
        if len(barcode) == 14 and barcode.startswith("0"):
            candidates.append(barcode[1:])
            
        # Candidate 5: If barcode scanned was 13 digits starting with '0', try the 12 digit UPC-A
        if len(barcode) == 13 and barcode.startswith("0"):
            candidates.append(barcode[1:])
            
    return candidates


def _match(header, aliases):
    if not header:
        return False
    hu = str(header).strip().upper()
    return hu in [a.upper() for a in aliases]


def detect_pdt_columns(file_content, file_type="csv"):
    """
    Parses headers from uploaded file content and proposes a default column mapping.
    """
    if len(file_content) > 5 * 1024 * 1024:
        frappe.throw("PDT file too large. Maximum 5MB allowed.")
        
    headers = []
    mapping = {k: "" for k in COL_ALIASES.keys()}
    
    if file_type in ("csv", "tsv"):
        delimiter = "\t" if file_type == "tsv" else ","
        content_str = file_content.decode("utf-8", errors="ignore") if isinstance(file_content, bytes) else str(file_content)
        lines = content_str.splitlines()
        if lines:
            reader = csv.reader(io.StringIO(lines[0]), delimiter=delimiter)
            headers = next(reader, [])
            
    elif file_type == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(max_row=1, values_only=True):
            headers = [str(val or "").strip() for val in row]
            break
            
    # Propose default mapping
    for h in headers:
        hu = h.strip().upper()
        for key, aliases in COL_ALIASES.items():
            if any(_match(hu, [a]) for a in aliases):
                mapping[key] = h
                break
                
    return headers, mapping


def resolve_barcode(barcode):
    """
    Barcode se item details resolve karo.
    Priority: Item Barcode table → Direct Item name match
    Also falls back to template settings for variants if fields are empty/unset.
    """
    frappe.has_permission("Item", "read", throw=True)
    
    if not barcode:
        return {"error": "Empty barcode", "barcode": ""}
    
    candidates = get_barcode_candidates(barcode)
    
    item_code = None
    for cand in candidates:
        item_code = frappe.db.get_value(
            "Item Barcode", {"barcode": cand}, "parent"
        )
        if item_code:
            break
            
    if not item_code:
        for cand in candidates:
            if frappe.db.exists("Item", cand):
                item_code = cand
                break
            
    if not item_code:
        return {"error": "Barcode not found", "barcode": barcode}
    
    item = frappe.get_doc("Item", item_code)
    article = item.variant_of or item.name
    category = item.item_group or ""
    
    # MRP fallback logic (variant -> parent template -> standard_rate)
    mrp = flt(item.get("custom_mrp"))
    if not mrp and item.variant_of:
        mrp = flt(frappe.db.get_value("Item", item.variant_of, "custom_mrp"))
    if not mrp:
        mrp = flt(item.valuation_rate or item.standard_rate or 0)
        
    # GST fallback logic (variant -> parent template -> 12.0)
    gst_pct = flt(item.get("custom_gst_percentage"))
    if not gst_pct and item.variant_of:
        gst_pct = flt(frappe.db.get_value("Item", item.variant_of, "custom_gst_percentage"))
    if not gst_pct:
        gst_pct = 12.0
        
    # HSN fallback logic (variant -> parent template)
    hsn_code = item.gst_hsn_code
    if not hsn_code and item.variant_of:
        hsn_code = frappe.db.get_value("Item", item.variant_of, "gst_hsn_code")
        
    # Sub-category fallback logic (variant -> parent template)
    sub_category = item.get("custom_sub_category")
    if not sub_category and item.variant_of:
        sub_category = frappe.db.get_value("Item", item.variant_of, "custom_sub_category")
        
    # Extract Color and Size from attributes
    attributes = {a.attribute.strip().upper(): a.attribute_value for a in item.attributes if a.attribute_value}
    color = attributes.get("COLOR") or attributes.get("COLOUR") or ""
    size = attributes.get("SIZE") or ""
    
    # Fallback to splitting item_code if color or size is still missing
    if item.variant_of and (not color or not size):
        parts = item_code.replace(article, "", 1).lstrip("-").split("-")
        if not color and len(parts) >= 1:
            color = parts[0]
        if not size and len(parts) >= 2:
            size = parts[-1]
            
    # Calculate B2B rate (excluding GST)
    rate = flt(mrp / (1 + (gst_pct / 100.0)), 2) if mrp > 0 else 0.0
    
    return {
        "item_code":    item_code,
        "item_name":    item.item_name,
        "stock_uom":    item.stock_uom,
        "brand":        item.brand,
        "item_group":   item.item_group,
        "article":      article,
        "color":        color,
        "size":         size,
        "mrp":          mrp,
        "rate":         rate,
        "gst_pct":      gst_pct,
        "hsn_code":     hsn_code or "",
        "category":     category,
        "sub_category": sub_category or "",
    }


def add_or_merge_item(rows, item_code, qty, rate, discount=0, **extra):
    """
    Shared merge logic.
    Same item_code + same rate → qty merge
    Same item_code + different rate → new separate row
    """
    rate  = round(flt(rate), 2)
    qty   = flt(qty)
    
    for row in rows:
        if (str(row.get("item_code")) == str(item_code) and 
                round(flt(row.get("rate")), 2) == rate):
            row["qty"] = flt(row.get("qty", 0)) + qty
            return rows, "merged"
    
    # New row
    new_row = {
        "item_code": item_code,
        "qty":       qty,
        "rate":      rate,
        "discount":  discount,
        **extra
    }
    rows.append(new_row)
    return rows, "added"


def parse_pdt_file(file_content, file_type="csv", col_mapping=None, price_type="Selling", supplier=None):
    """
    PDT file parse karo — CSV/TSV/Excel support.
    
    Only BARCODE is required.
    Missing fields fallback logic:
      - Price missing -> Item Master custom_mrp
      - Discount missing -> 0
      - Qty missing -> 1
      - Tax missing -> Item Master custom_gst_percentage
    """
    if len(file_content) > 5 * 1024 * 1024:
        frappe.throw("PDT file too large. Maximum 5MB allowed.")
        
    # Propose default mapping if not provided
    if not col_mapping:
        _, col_mapping = detect_pdt_columns(file_content, file_type)
        
    col_mapping = col_mapping or {}
    
    raw_rows = []
    
    if file_type in ("csv", "tsv"):
        delimiter = "\t" if file_type == "tsv" else ","
        content_str = file_content.decode("utf-8", errors="ignore") if isinstance(file_content, bytes) else str(file_content)
        reader = csv.DictReader(io.StringIO(content_str), delimiter=delimiter)
        for row in reader:
            raw_rows.append(row)
            
    elif file_type == "xlsx":
        import openpyxl
        if isinstance(file_content, str):
            file_content = file_content.encode("utf-8", errors="ignore")
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        for r_idx in range(2, sheet.max_row + 1):
            row_cells = sheet[r_idx]
            row = {}
            for h_idx, h in enumerate(headers):
                if h_idx < len(row_cells):
                    row[h] = row_cells[h_idx].value
            raw_rows.append(row)
            
    parsed_rows = []
    
    for row in raw_rows:
        # Get barcode
        barcode_key = col_mapping.get("barcode", "")
        barcode = str(row.get(barcode_key, "") or "").strip()
        
        # Skip empty/blank rows
        if not barcode or barcode.lower() in ("", "nan", "none", "null", "0"):
            continue
            
        # Get raw values
        price_key = col_mapping.get("price", "")
        disc_key = col_mapping.get("discount", "")
        qty_key = col_mapping.get("qty", "")
        tax_key = col_mapping.get("tax", "")
        
        raw_price = row.get(price_key) if price_key else None
        raw_disc = row.get(disc_key) if disc_key else None
        raw_qty = row.get(qty_key) if qty_key else None
        raw_tax = row.get(tax_key) if tax_key else None
        
        # 1. Price Source
        if price_key and raw_price is not None and str(raw_price).strip() != "":
            price = flt(raw_price)
            price_src = "pdt"
        else:
            price = None
            price_src = "item_master"
            
        # 2. Discount Source
        if disc_key and raw_disc is not None and str(raw_disc).strip() != "":
            discount = flt(raw_disc)
            disc_src = "pdt"
        else:
            discount = 0.0
            disc_src = "default"
            
        # 3. Qty Source
        if qty_key and raw_qty is not None and str(raw_qty).strip() != "":
            qty = flt(raw_qty) or 1.0
            qty_src = "pdt"
        else:
            qty = 1.0
            qty_src = "default"
            
        # 4. Tax Source
        if tax_key and raw_tax is not None and str(raw_tax).strip() != "":
            tax = flt(raw_tax)
            tax_src = "pdt"
        else:
            tax = None
            tax_src = "item_master"
            
        # Resolve barcode details for item master fallback
        item_details = {}
        error = None
        
        # We need item details if barcode is queried or for preview formatting
        try:
            res = resolve_barcode(barcode)
            if isinstance(res, dict) and "error" in res:
                error = res["error"]
                if price_src == "item_master":
                    price_src = "not_found"
                if tax_src == "item_master":
                    tax_src = "not_found"
            else:
                item_details = res
                if price_src == "item_master":
                    if price_type == "Buying":
                        price = get_buying_rate(item_details.get("item_code"), supplier)
                    else:
                        price = flt(item_details.get("mrp") or 0.0)
                if tax_src == "item_master":
                    tax = flt(item_details.get("gst_pct") or 12.0)
        except Exception as e:
            error = str(e)
            if price_src == "item_master":
                price_src = "not_found"
            if tax_src == "item_master":
                tax_src = "not_found"
                
        # Calculated B2B rate
        resolved_rate = flt(price / (1 + ((tax or 12.0) / 100.0)), 2) if price and price > 0 else 0.0
        
        parsed_row = {
            "barcode": barcode,
            "price": price,
            "discount": discount,
            "qty": qty,
            "tax": tax,
            "rate": resolved_rate,
            "_price_source": price_src,
            "_discount_source": disc_src,
            "_qty_source": qty_src,
            "_tax_source": tax_src,
        }
        
        if error:
            parsed_row["error"] = error
        else:
            # Append item details for frontend grid rendering
            parsed_row.update({
                "item_code": item_details.get("item_code"),
                "item_name": item_details.get("item_name"),
                "stock_uom": item_details.get("stock_uom"),
                "brand": item_details.get("brand"),
                "item_group": item_details.get("item_group"),
                "article": item_details.get("article"),
                "color": item_details.get("color"),
                "size": item_details.get("size"),
                "hsn_code": item_details.get("hsn_code"),
                "category": item_details.get("category"),
                "sub_category": item_details.get("sub_category"),
            })
            
        parsed_rows.append(parsed_row)
        
    return parsed_rows


def get_buying_rate(item_code, supplier=None):
    """
    Resolves supplier-specific buying rate with priority:
    1. Supplier default_price_list -> Item Price
    2. "Standard Buying" price list -> Item Price
    3. Item.valuation_rate
    4. Item.standard_rate
    5. 0.0 (last resort)
    """
    if not item_code:
        return 0.0

    price_list = None
    if supplier:
        price_list = frappe.db.get_value("Supplier", supplier, "default_price_list")
        
    if not price_list:
        price_list = "Standard Buying"

    # 1 & 2: Query Item Price matching price list
    rate = frappe.db.get_value(
        "Item Price", 
        {"item_code": item_code, "price_list": price_list}, 
        "price_list_rate"
    )
    if not rate and price_list != "Standard Buying":
        rate = frappe.db.get_value(
            "Item Price", 
            {"item_code": item_code, "price_list": "Standard Buying"}, 
            "price_list_rate"
        )

    # 3 & 4: Fallback to Item valuation_rate / standard_rate
    if not rate:
        item = frappe.get_doc("Item", item_code)
        rate = item.valuation_rate or item.standard_rate or 0.0

    return flt(rate)
