# -*- coding: utf-8 -*-
#
# @file: smriti_retail_os/setup_demo_company.py
# @description: Demo company provisioning — Excel-driven company, address, bank, and logo setup.
# @author: Jawahar R Mallah <jawahar.mallah@gmail.com>
# @date: 2026-05-28
# @version: 1.8.6
# @license: MIT
# * Copyright (c) 2026 AITDL NETWORK & ERPNbook.com. All rights reserved.
#
import frappe
import openpyxl
import os

def parse_excel_setup(file_path=None):
    if not file_path:
        # Search for any setup template Excel in the company directory
        company_dir = "/home/frappe/frappe-bench/company"
        if os.path.exists(company_dir):
            for f in os.listdir(company_dir):
                if f.endswith(".xlsx") and "template" in f.lower():
                    file_path = os.path.join(company_dir, f)
                    break
            if not file_path:
                # Fallback: look for any xlsx
                for f in os.listdir(company_dir):
                    if f.endswith(".xlsx"):
                        file_path = os.path.join(company_dir, f)
                        break
    if not file_path or not os.path.exists(file_path):
        print(f"ERROR: Excel setup file not found at {file_path}")
        return None
        
    print(f"Parsing Excel setup sheet: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    data = {}
    
    # ── 1. Company Master (Col 1 for key, Col 3 for value) ────────────────────
    sheet1 = wb["1. Company Master"]
    for row in sheet1.iter_rows(values_only=True):
        if len(row) >= 3 and row[0]:
            field = str(row[0]).strip().replace("*", "").strip().lower()
            val = str(row[2]).strip() if row[2] is not None else ""
            if field == "company name": data["company_name"] = val
            elif field == "abbr": data["abbr"] = val
            elif field == "company pan": data["pan"] = val
            elif field == "phone": data["phone_no"] = val
            elif field == "email": data["email"] = val
            elif field == "address line 1": data["address_line1"] = val
            elif field == "address line 2": data["address_line2"] = val
            elif field == "city": data["city"] = val
            elif field == "pin code": data["pincode"] = val
            elif field == "state": data["state"] = val
            elif field == "country": data["country"] = val
            
    # ── 2. GST Configuration (Col 1 for key, Col 3 for value) ─────────────────
    sheet2 = wb["2. GST Configuration"]
    for row in sheet2.iter_rows(values_only=True):
        if len(row) >= 3 and row[0]:
            field = str(row[0]).strip().replace("*", "").strip().lower()
            val = str(row[2]).strip() if row[2] is not None else ""
            if field == "gstin / uin": data["gstin"] = val
            elif field == "state code": data["state_code"] = val
            
    # ── 4. Bank Account Ledger (Col 1 for key, Col 2 for value) ───────────────
    sheet4 = wb["4. Bank Account Ledger"]
    for row in sheet4.iter_rows(values_only=True):
        if len(row) >= 2 and row[0]:
            field = str(row[0]).strip().replace("*", "").strip().lower()
            val = str(row[1]).strip() if row[1] is not None else ""
            if field == "bank name": data["bank_name"] = val
            elif field == "branch": data["branch"] = val
            elif field == "account number": data["bank_account_no"] = val
            elif field == "ifsc code": data["branch_code"] = val # IFSC
            
    return data

def run():
    print("Starting Demo Company Dynamic Programmatic Setup...")
    
    # 0. Pre-create standard Warehouse Types to avoid LinkValidationErrors
    warehouse_types = ["All", "Transit", "Bonded", "Consignment", "Spares", "Work In Progress"]
    for wt_name in warehouse_types:
        if not frappe.db.exists("Warehouse Type", wt_name):
            try:
                wt = frappe.new_doc("Warehouse Type")
                wt.name = wt_name
                wt.warehouse_type = wt_name
                wt.insert(ignore_permissions=True)
                frappe.db.commit()
                print(f"Created standard Warehouse Type: {wt_name}")
            except Exception as e:
                print(f"Error creating Warehouse Type {wt_name}: {e}")
                
    # 1. Parse Excel data
    info = parse_excel_setup()
    if not info:
        print("Aborting setup: could not parse Excel setup.")
        return
        
    company_name = info.get("company_name")
    if not company_name:
        print("ERROR: company_name not found in Excel setup sheet.")
        return
    abbr = info.get("abbr", "")
    
    print(f"Retrieved Company Name: '{company_name}' from Excel.")
    
    # 2. Upsert Company
    logo_filename = info.get("logo_filename", "company_logo.png")
    src_logo = os.path.join("/home/frappe/frappe-bench/company", logo_filename)
    company_logo_path = None
    if os.path.exists(src_logo):
        try:
            import shutil
            dest_dir = frappe.get_site_path("public", "files")
            if not os.path.exists(dest_dir):
                os.makedirs(dest_dir, exist_ok=True)
            dest_logo = os.path.join(dest_dir, logo_filename)
            shutil.copy(src_logo, dest_logo)
            # Ensure permissions are correct
            os.chmod(dest_logo, 0o644)
            company_logo_path = f"/files/{logo_filename}"
            print(f"Successfully copied company logo to {dest_logo}")
        except Exception as e:
            print(f"Error copying company logo: {e}")

    if not frappe.db.exists("Company", company_name):
        co = frappe.new_doc("Company")
        co.company_name = company_name
        co.abbr = abbr
        co.default_currency = info.get("default_currency", "INR")
        co.country = info.get("country", "India")
        co.phone_no = info.get("phone_no")
        co.email = info.get("email")
        co.pan = info.get("pan")
        co.tax_id = info.get("gstin")
        co.gstin = info.get("gstin")
        co.domain = "Retail"
        if company_logo_path:
            co.company_logo = company_logo_path
        co.insert(ignore_permissions=True)
        frappe.db.commit()
        print(f"SUCCESS: Created Company '{company_name}'!")
    else:
        co = frappe.get_doc("Company", company_name)
        co.phone_no = info.get("phone_no")
        co.email = info.get("email")
        co.pan = info.get("pan")
        co.tax_id = info.get("gstin")
        co.gstin = info.get("gstin")
        if company_logo_path:
            co.company_logo = company_logo_path
        co.save(ignore_permissions=True)
        frappe.db.commit()
        print(f"SUCCESS: Updated Company '{company_name}' details!")

    # 3. Set as Default Company
    frappe.defaults.set_global_default("company", company_name)
    frappe.defaults.set_user_default("company", company_name, "Administrator")
    print(f"Set '{company_name}' as the global default company.")

    # 4. Create/Update Registered Office Address
    address_name = f"{company_name}-Registered"
    if not frappe.db.exists("Address", address_name):
        addr = frappe.new_doc("Address")
        addr.address_title = company_name
        addr.address_type = "Office"
        addr.address_line1 = info.get("address_line1")
        addr.address_line2 = info.get("address_line2")
        addr.city = info.get("city")
        addr.state = info.get("state")
        addr.pincode = info.get("pincode")
        addr.country = info.get("country", "India")
        addr.email_id = info.get("email")
        addr.phone = info.get("phone_no")
        addr.is_primary_address = 1
        addr.is_shipping_address = 1
        addr.is_your_company_address = 1
        addr.gstin = info.get("gstin")
        addr.gst_state = info.get("state")
        addr.gst_state_number = info.get("state_code")
        addr.gst_category = "Registered"
        addr.append("links", {"link_doctype": "Company", "link_name": company_name})
        addr.insert(ignore_permissions=True)
        frappe.db.commit()
        print("Registered Address created and linked to Company!")
    else:
        addr = frappe.get_doc("Address", address_name)
        addr.address_line1 = info.get("address_line1")
        addr.address_line2 = info.get("address_line2")
        addr.city = info.get("city")
        addr.state = info.get("state")
        addr.pincode = info.get("pincode")
        addr.gstin = info.get("gstin")
        addr.gst_state = info.get("state")
        addr.gst_state_number = info.get("state_code")
        addr.email_id = info.get("email")
        addr.phone = info.get("phone_no")
        addr.save(ignore_permissions=True)
        frappe.db.commit()
        print("Registered Address successfully updated to match Excel changes!")

    # 5. Create Bank Account and Bank ledger
    bank_name = info.get("bank_name", "State Bank of India")
    if bank_name:
        if not frappe.db.exists("Bank", bank_name):
            b = frappe.new_doc("Bank")
            b.bank_name = bank_name
            b.insert(ignore_permissions=True)
            frappe.db.commit()
            print(f"Bank '{bank_name}' created.")

        # Create Bank ledger under Bank Accounts
        parent_account = f"Bank Accounts - {co.abbr}"
        
        if frappe.db.exists("Account", parent_account):
            ledger_name = f"{bank_name} - {co.abbr}"
            if not frappe.db.exists("Account", ledger_name):
                acc = frappe.new_doc("Account")
                acc.account_name = bank_name
                acc.parent_account = parent_account
                acc.company = company_name
                acc.account_type = "Bank"
                acc.insert(ignore_permissions=True)
                frappe.db.commit()
                print(f"Bank ledger '{acc.name}' created.")
                account_id = acc.name
            else:
                account_id = ledger_name
                print(f"Bank ledger '{ledger_name}' already exists.")
        else:
            # Fallback creation
            assets_parent = f"Current Assets - {co.abbr}"
            if frappe.db.exists("Account", assets_parent):
                group_acc = frappe.new_doc("Account")
                group_acc.account_name = "Bank Accounts"
                group_acc.parent_account = assets_parent
                group_acc.company = company_name
                group_acc.is_group = 1
                group_acc.insert(ignore_permissions=True)
                frappe.db.commit()
                parent_account = group_acc.name
                
                acc = frappe.new_doc("Account")
                acc.account_name = bank_name
                acc.parent_account = parent_account
                acc.company = company_name
                acc.account_type = "Bank"
                acc.insert(ignore_permissions=True)
                frappe.db.commit()
                print(f"Bank ledger '{acc.name}' created.")
                account_id = acc.name
            else:
                account_id = None

        # Create Bank Account doc
        bank_account_no = info.get("bank_account_no")
        if bank_account_no:
            existing_ba = frappe.db.get_value("Bank Account", {"bank_account_no": bank_account_no}, "name")
            if not existing_ba:
                ba = frappe.new_doc("Bank Account")
                ba.account_name = bank_name
                ba.bank = bank_name
                ba.bank_account_no = bank_account_no
                ba.company = company_name
                ba.is_company_account = 1
                ba.is_default = 1
                ba.branch_code = info.get("branch_code")
                if account_id:
                    ba.account = account_id
                ba.insert(ignore_permissions=True)
                frappe.db.commit()
                print(f"Bank Account '{ba.name}' created successfully!")
            else:
                ba = frappe.get_doc("Bank Account", existing_ba)
                ba.bank_account_no = bank_account_no
                ba.branch_code = info.get("branch_code")
                if account_id:
                    ba.account = account_id
                ba.save(ignore_permissions=True)
                frappe.db.commit()
                print(f"Bank Account '{existing_ba}' successfully updated to match Excel changes!")
        
    # 6. Update Website Settings & Navbar Settings with the logo
    if company_logo_path:
        try:
            ws = frappe.get_doc("Website Settings")
            ws.app_logo = company_logo_path
            ws.banner_image = company_logo_path
            ws.save(ignore_permissions=True)
            frappe.db.commit()
            print("Successfully updated Website Settings with company logo.")
        except Exception as e:
            print(f"Error updating Website Settings: {e}")

        try:
            ns = frappe.get_doc("Navbar Settings")
            ns.app_logo = company_logo_path
            ns.save(ignore_permissions=True)
            frappe.db.commit()
            print("Successfully updated Navbar Settings with company logo.")
        except Exception as e:
            print(f"Error updating Navbar Settings: {e}")

    print("Demo Company Dynamic Programmatic Setup finished successfully!")
